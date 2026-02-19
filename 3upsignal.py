import pandas as pd
import numpy as np
import warnings
import os
from datetime import datetime

# === Setup ===
warnings.simplefilter(action="ignore", category=FutureWarning)
os.makedirs("output", exist_ok=True)

# === User Configuration ===
POSITION_SIZE = 15000  # Fixed position size per trade
DATA_START_DATE = "2023-09-01"  # Load data from here (for signal calculation)
PORTFOLIO_START_DATE = "2024-09-04"  # Start portfolio trading from here
END_DATE = "2025-09-05"

# === Define the 76 tickers from your trading ===
TICKERS = [
    "ALAB", "AMD", "AMZN", "ANET", "APP", "ARKK", "AS", "ASAN", "AXON", 
    "BIRK", "BKNG", "BROS", "CAVA", "CCL", "CHYM", "COIN", "COST", "CRCL", 
    "CRDO", "CRWV", "DASH", "DECK", "DHI", "DOCS", "DUOL", "ETHU", "ETOR", 
    "FETH", "FIX", "FNGO", "GDX", "GEV", "GOOGL", "GRMN", "GTLB", "HIMS", 
    "HOOD", "IBIT", "IBKR", "IOT", "IREN", "ISRG", "KVYO", "LBTR", "LULU", 
    "META", "MGNI", "MSTR", "NBIS", "NET", "NFLX", "NRG", "NVDA", "ODD", 
    "OKLO", "PLTR", "QQQ", "RBLX", "RBRK", "RDDT", "RKLB", "SE", "SHOP", 
    "SNOW", "SPOT", "TOST", "TQQQ", "TSLA", "TWLO", "UAL", "UBER", "UPRO", 
    "VIK", "XHB", "XYZ", "ZETA"
]

print(f"📊 3 Up Signal Strategy Tester")
print(f"💰 Position Size: ${POSITION_SIZE:,.2f}")
print(f"📅 Test Period: {PORTFOLIO_START_DATE} to {END_DATE}")
print(f"🎯 Testing {len(TICKERS)} tickers\n")

# === Load and prepare all ticker data ===
print("🔄 Loading price data...")
ticker_data = {}
missing_tickers = []
all_dates = set()

for ticker in TICKERS:
    input_file = f"output/{ticker}_price_data.csv"
    
    try:
        df = pd.read_csv(input_file)
    except FileNotFoundError:
        missing_tickers.append(ticker)
        continue
    
    # Filter date range
    df["Date"] = pd.to_datetime(df["Date"])
    df.set_index("Date", inplace=True)
    df = df[(df.index >= DATA_START_DATE) & (df.index <= END_DATE)]
    
    if df.empty:
        print(f"⚠️ No data in date range for {ticker}")
        continue
    
    # Clean data
    for col in ["Open", "High", "Low", "Close"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df.dropna(subset=["Open", "High", "Low", "Close"], inplace=True)
    
    # Calculate EMA21
    df["EMA21"] = df["Close"].ewm(span=21, adjust=False).mean()
    
    # Generate signals using 3 Up Signal logic
    df["signal"] = 0
    position_open = False
    first_break_low = None
    
    for i in range(2, len(df)):
        # BUY SIGNAL: 3 consecutive days with low > EMA21 AND current day is up day
        if not position_open and i >= 2:
            three_days_above = (df.iloc[i]["Low"] > df.iloc[i]["EMA21"] and 
                               df.iloc[i-1]["Low"] > df.iloc[i-1]["EMA21"] and 
                               df.iloc[i-2]["Low"] > df.iloc[i-2]["EMA21"])
            is_up_day = df.iloc[i]["Close"] > df.iloc[i-1]["Close"]
            
            if three_days_above and is_up_day:
                df.at[df.index[i], "signal"] = 1
                position_open = True
                first_break_low = None
        
        # SELL SIGNAL: Two-step process
        elif position_open:
            # Step 1: Check if we closed below EMA for first time
            if first_break_low is None:
                if df.iloc[i]["Close"] < df.iloc[i]["EMA21"]:
                    if i > 0 and df.iloc[i-1]["Close"] >= df.iloc[i-1]["EMA21"]:
                        first_break_low = df.iloc[i]["Low"]
            
            # Step 2: Sell if low drops below the stored low
            if first_break_low is not None:
                if df.iloc[i]["Low"] < first_break_low:
                    df.at[df.index[i], "signal"] = -1
                    position_open = False
                    first_break_low = None
    
    ticker_data[ticker] = df
    all_dates.update(df.index)

# Report missing tickers
if missing_tickers:
    print(f"\n⚠️ Missing price data for {len(missing_tickers)} tickers:")
    with open("output/missing_tickers.txt", "w") as f:
        for ticker in missing_tickers:
            f.write(f"{ticker}\n")
            print(f"   - {ticker}")
    print(f"💾 Missing ticker list saved to: output/missing_tickers.txt")
    print("📌 Add these to select_ticker.txt and run the scraper\n")

print(f"✅ Successfully loaded data for {len(ticker_data)} tickers")

# Sort dates and filter for portfolio period
all_dates = sorted(list(all_dates))
portfolio_dates = [d for d in all_dates if d >= pd.to_datetime(PORTFOLIO_START_DATE)]

# === Simulate 3 Up Signal Strategy ===
print("\n🔄 Running 3 Up Signal strategy simulation...")

all_trades = []
trade_num = 0

for ticker in ticker_data:
    df = ticker_data[ticker]
    position_open = False
    entry_date = None
    entry_price = None
    
    # Check for pre-existing position before portfolio start
    for date in df.index:
        if date < pd.to_datetime(PORTFOLIO_START_DATE):
            if df.loc[date, "signal"] == 1:
                position_open = True
                entry_date = date
                entry_price = df.loc[date, "Close"]
            elif df.loc[date, "signal"] == -1:
                position_open = False
                entry_date = None
                entry_price = None
        else:
            break
    
    # Process signals during portfolio period
    for date in df.index:
        if date < pd.to_datetime(PORTFOLIO_START_DATE):
            continue
        if date > pd.to_datetime(END_DATE):
            break
        
        # Buy signal
        if not position_open and df.loc[date, "signal"] == 1:
            entry_date = date
            entry_price = df.loc[date, "Close"]
            position_open = True
        
        # Sell signal
        elif position_open and df.loc[date, "signal"] == -1:
            exit_date = date
            exit_price = df.loc[date, "Close"]
            
            # Only record if entry was during portfolio period
            if entry_date >= pd.to_datetime(PORTFOLIO_START_DATE):
                trade_num += 1
                shares = POSITION_SIZE / entry_price
                exit_value = shares * exit_price
                pnl_dollars = exit_value - POSITION_SIZE
                pnl_pct = ((exit_price - entry_price) / entry_price) * 100
                
                all_trades.append({
                    'Trade_Num': trade_num,
                    'Ticker': ticker,
                    'Entry_Date': entry_date.strftime('%Y-%m-%d'),
                    'Exit_Date': exit_date.strftime('%Y-%m-%d'),
                    'Entry_Price': round(entry_price, 2),
                    'Exit_Price': round(exit_price, 2),
                    'Shares': round(shares, 2),
                    'Position_Size': POSITION_SIZE,
                    'Exit_Value': round(exit_value, 2),
                    'PnL_Dollars': round(pnl_dollars, 2),
                    'PnL_Pct': round(pnl_pct, 2),
                    'Status': 'Closed'
                })
            
            position_open = False
            entry_date = None
            entry_price = None
    
    # Handle open positions at end of period
    if position_open and entry_date >= pd.to_datetime(PORTFOLIO_START_DATE):
        last_date = min(pd.to_datetime(END_DATE), df.index[-1])
        if last_date in df.index:
            trade_num += 1
            exit_price = df.loc[last_date, "Close"]
            shares = POSITION_SIZE / entry_price
            exit_value = shares * exit_price
            pnl_dollars = exit_value - POSITION_SIZE
            pnl_pct = ((exit_price - entry_price) / entry_price) * 100
            
            all_trades.append({
                'Trade_Num': trade_num,
                'Ticker': ticker,
                'Entry_Date': entry_date.strftime('%Y-%m-%d'),
                'Exit_Date': last_date.strftime('%Y-%m-%d'),
                'Entry_Price': round(entry_price, 2),
                'Exit_Price': round(exit_price, 2),
                'Shares': round(shares, 2),
                'Position_Size': POSITION_SIZE,
                'Exit_Value': round(exit_value, 2),
                'PnL_Dollars': round(pnl_dollars, 2),
                'PnL_Pct': round(pnl_pct, 2),
                'Status': 'Open (Hypothetical Close)'
            })

print(f"✅ Generated {len(all_trades)} trades")

# === Calculate Statistics ===
if all_trades:
    trades_df = pd.DataFrame(all_trades)
    
    # Calculate summary statistics
    winning_trades = trades_df[trades_df['PnL_Dollars'] > 0]
    losing_trades = trades_df[trades_df['PnL_Dollars'] < 0]
    closed_trades = trades_df[trades_df['Status'] == 'Closed']
    open_trades = trades_df[trades_df['Status'] != 'Closed']
    
    total_pnl = trades_df['PnL_Dollars'].sum()
    win_rate = (len(winning_trades) / len(trades_df)) * 100 if len(trades_df) > 0 else 0
    
    # === Save Results to Excel ===
    with pd.ExcelWriter("output/3up_signal_results.xlsx", engine='openpyxl') as writer:
        # Detailed Trades Sheet
        trades_df.to_excel(writer, sheet_name='All Trades', index=False)
        
        # Summary Statistics Sheet
        summary_stats = {
            'Metric': [
                'Total Trades',
                'Closed Trades',
                'Open Trades',
                'Winning Trades',
                'Losing Trades',
                'Win Rate (%)',
                'Total P&L ($)',
                'Total Gains ($)',
                'Total Losses ($)',
                'Average Win ($)',
                'Average Loss ($)',
                'Largest Win ($)',
                'Largest Loss ($)',
                'Average Trade P&L ($)',
                'Unique Tickers Traded',
                'Position Size ($)'
            ],
            'Value': [
                len(trades_df),
                len(closed_trades),
                len(open_trades),
                len(winning_trades),
                len(losing_trades),
                round(win_rate, 2),
                round(total_pnl, 2),
                round(winning_trades['PnL_Dollars'].sum(), 2) if not winning_trades.empty else 0,
                round(losing_trades['PnL_Dollars'].sum(), 2) if not losing_trades.empty else 0,
                round(winning_trades['PnL_Dollars'].mean(), 2) if not winning_trades.empty else 0,
                round(losing_trades['PnL_Dollars'].mean(), 2) if not losing_trades.empty else 0,
                round(winning_trades['PnL_Dollars'].max(), 2) if not winning_trades.empty else 0,
                round(losing_trades['PnL_Dollars'].min(), 2) if not losing_trades.empty else 0,
                round(trades_df['PnL_Dollars'].mean(), 2),
                trades_df['Ticker'].nunique(),
                POSITION_SIZE
            ]
        }
        
        summary_df = pd.DataFrame(summary_stats)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Ticker Performance Sheet
        ticker_summary = trades_df.groupby('Ticker').agg({
            'Trade_Num': 'count',
            'PnL_Dollars': 'sum',
            'PnL_Pct': 'mean'
        }).round(2)
        ticker_summary.columns = ['Trades', 'Total_PnL', 'Avg_PnL_Pct']
        ticker_summary = ticker_summary.sort_values('Total_PnL', ascending=False)
        ticker_summary.to_excel(writer, sheet_name='By Ticker')
        
        # Format the Excel file
        workbook = writer.book
        
        # Format Summary sheet
        worksheet = writer.sheets['Summary']
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Format headers
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Format P&L row
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            if row[0].value == 'Total P&L ($)':
                if row[1].value >= 0:
                    row[1].font = Font(color="008000", bold=True, size=12)
                else:
                    row[1].font = Font(color="FF0000", bold=True, size=12)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"\n💾 Results saved to: output/3up_signal_results.xlsx")
    
    # === Print Summary to Console ===
    print("\n" + "="*60)
    print("📊 3 UP SIGNAL STRATEGY RESULTS")
    print("="*60)
    print(f"Test Period: {PORTFOLIO_START_DATE} to {END_DATE}")
    print(f"Position Size: ${POSITION_SIZE:,.2f}")
    print(f"Tickers Analyzed: {len(ticker_data)}")
    print(f"Tickers with Trades: {trades_df['Ticker'].nunique()}")
    print()
    print(f"Total Trades: {len(trades_df)}")
    print(f"  - Closed: {len(closed_trades)}")
    print(f"  - Open: {len(open_trades)}")
    print()
    print(f"Win Rate: {win_rate:.1f}%")
    print(f"  - Winners: {len(winning_trades)}")
    print(f"  - Losers: {len(losing_trades)}")
    print()
    print(f"P&L Summary:")
    print(f"  - Total P&L: ${total_pnl:,.2f}")
    print(f"  - Total Gains: ${winning_trades['PnL_Dollars'].sum():,.2f}" if not winning_trades.empty else "  - Total Gains: $0.00")
    print(f"  - Total Losses: ${losing_trades['PnL_Dollars'].sum():,.2f}" if not losing_trades.empty else "  - Total Losses: $0.00")
    print(f"  - Average Win: ${winning_trades['PnL_Dollars'].mean():,.2f}" if not winning_trades.empty else "  - Average Win: $0.00")
    print(f"  - Average Loss: ${losing_trades['PnL_Dollars'].mean():,.2f}" if not losing_trades.empty else "  - Average Loss: $0.00")
    print()
    print(f"Top 5 Performers:")
    top_tickers = trades_df.groupby('Ticker')['PnL_Dollars'].sum().sort_values(ascending=False).head(5)
    for ticker, pnl in top_tickers.items():
        print(f"  - {ticker}: ${pnl:,.2f}")
    print()
    print(f"Bottom 5 Performers:")
    bottom_tickers = trades_df.groupby('Ticker')['PnL_Dollars'].sum().sort_values().head(5)
    for ticker, pnl in bottom_tickers.items():
        print(f"  - {ticker}: ${pnl:,.2f}")
    
else:
    print("\n⚠️ No trades generated. Check if:")
    print("  1. Price data files exist for the tickers")
    print("  2. The date range contains valid trading days")
    print("  3. The 3 Up Signal conditions were met")

print("\n✅ Analysis complete!")