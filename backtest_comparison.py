import pandas as pd
import numpy as np
import warnings
import os
from datetime import datetime
import random

# === Setup ===
warnings.simplefilter(action="ignore", category=FutureWarning)
os.makedirs("output", exist_ok=True)
random.seed(42)  # For reproducibility

# === User Configuration ===
INITIAL_CAPITAL = 150000  # Starting portfolio value
MAX_LEVERAGE = 1.8  # Can invest up to 180% of portfolio value
POSITION_SIZE_PCT = 0.10  # 10% of NLV per position
DATA_START_DATE = "2023-09-01"  # Load data from here (for signal calculation)
PORTFOLIO_START_DATE = "2024-09-04"  # Start portfolio trading from here
END_DATE = "2025-09-05"

# === Tickers list (based on uploaded files) ===
tickers = ["APP", "CRDO", "AMD", "NVDA", "RDDT", "PLTR", "HOOD", "GEV", "CRWV", "ALAB"]

print(f"📊 PORTFOLIO BACKTEST COMPARISON")
print(f"💰 Initial Capital: ${INITIAL_CAPITAL:,.2f}")
print(f"📈 Max Leverage: {MAX_LEVERAGE:.1%}")
print(f"📦 Position Size: {POSITION_SIZE_PCT:.1%} of NLV")
print(f"📅 Data Period: {DATA_START_DATE} to {END_DATE}")
print(f"💼 Portfolio Start: {PORTFOLIO_START_DATE}")
print(f"🎯 Tickers: {len(tickers)}\n")
print("="*80)

# === Function to load and prepare ticker data ===
def load_ticker_data(tickers, data_start_date, end_date):
    ticker_data = {}
    all_dates = set()
    
    for ticker in tickers:
        input_file = f"output/{ticker}_price_data.csv"  # Changed to read from output folder
        
        try:
            df = pd.read_csv(input_file)
        except FileNotFoundError:
            print(f"⚠️ File not found: {input_file}. Skipping {ticker}.")
            continue
        
        # Filter date range
        df["Date"] = pd.to_datetime(df["Date"])
        df.set_index("Date", inplace=True)
        df = df[(df.index >= data_start_date) & (df.index <= end_date)]
        
        if df.empty:
            print(f"⚠️ No data in date range for {ticker}. Skipping.")
            continue
        
        # Clean data
        for col in ["Open", "High", "Low", "Close"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        df.dropna(subset=["Open", "High", "Low", "Close"], inplace=True)
        
        # Calculate EMA21
        df["EMA21"] = df["Close"].ewm(span=21, adjust=False).mean()
        
        ticker_data[ticker] = df
        all_dates.update(df.index)
    
    return ticker_data, sorted(list(all_dates))

# === Strategy 1: Original (Two-step exit) ===
def apply_original_strategy(ticker_data):
    for ticker in ticker_data:
        df = ticker_data[ticker]
        df["signal_original"] = 0
        position_open = False
        first_break_low = None
        
        for i in range(2, len(df)):
            # BUY SIGNAL: 3 consecutive days with low > EMA21 AND current day is up day
            if not position_open and i >= 2:
                three_days_above = (df.iloc[i]["Low"] > df.iloc[i]["EMA21"] and 
                                   df.iloc[i-1]["Low"] > df.iloc[i-1]["EMA21"] and 
                                   df.iloc[i-2]["Low"] > df.iloc[i-2]["EMA21"])
                is_up_day = df.iloc[i]["Close"] > df.iloc[i-1]["Close"]
                
                if three_days_above and is_up_day:
                    df.at[df.index[i], "signal_original"] = 1
                    position_open = True
                    first_break_low = None
            
            # SELL SIGNAL: Two-step process
            elif position_open:
                # Step 1: Check if we closed below EMA for first time
                if first_break_low is None:
                    if df.iloc[i]["Close"] < df.iloc[i]["EMA21"]:
                        if i > 0 and df.iloc[i-1]["Close"] >= df.iloc[i-1]["EMA21"]:
                            first_break_low = df.iloc[i]["Low"]
                
                # Step 2: Sell if low drops below the stored low
                if first_break_low is not None:
                    if df.iloc[i]["Low"] < first_break_low:
                        df.at[df.index[i], "signal_original"] = -1
                        position_open = False
                        first_break_low = None

# === Strategy 2: Exit after 2 closes below EMA21 ===
def apply_two_closes_strategy(ticker_data):
    for ticker in ticker_data:
        df = ticker_data[ticker]
        df["signal_2closes"] = 0
        position_open = False
        consecutive_closes_below = 0
        
        for i in range(2, len(df)):
            # BUY SIGNAL: Same as original - 3 consecutive days with low > EMA21 AND current day is up day
            if not position_open and i >= 2:
                three_days_above = (df.iloc[i]["Low"] > df.iloc[i]["EMA21"] and 
                                   df.iloc[i-1]["Low"] > df.iloc[i-1]["EMA21"] and 
                                   df.iloc[i-2]["Low"] > df.iloc[i-2]["EMA21"])
                is_up_day = df.iloc[i]["Close"] > df.iloc[i-1]["Close"]
                
                if three_days_above and is_up_day:
                    df.at[df.index[i], "signal_2closes"] = 1
                    position_open = True
                    consecutive_closes_below = 0
            
            # SELL SIGNAL: 2 consecutive closes below EMA21
            elif position_open:
                if df.iloc[i]["Close"] < df.iloc[i]["EMA21"]:
                    consecutive_closes_below += 1
                    if consecutive_closes_below >= 2:
                        df.at[df.index[i], "signal_2closes"] = -1
                        position_open = False
                        consecutive_closes_below = 0
                else:
                    consecutive_closes_below = 0  # Reset counter if close is above EMA

# === Portfolio Class ===
class Portfolio:
    def __init__(self, initial_capital, max_leverage, position_size_pct, strategy_name):
        self.initial_capital = initial_capital
        self.cash = initial_capital
        self.max_leverage = max_leverage
        self.position_size_pct = position_size_pct
        self.positions = {}
        self.trade_history = []
        self.pre_existing = set()
        self.strategy_name = strategy_name
        self.daily_values = []
        
    def get_nlv(self, date, ticker_data):
        nlv = self.cash
        for ticker, pos in self.positions.items():
            if ticker in ticker_data and date in ticker_data[ticker].index:
                current_price = ticker_data[ticker].loc[date, "Close"]
                nlv += pos['shares'] * current_price
        return nlv
    
    def get_buying_power(self, nlv):
        max_investment = nlv * self.max_leverage
        current_investment = sum(pos['shares'] * pos['entry_price'] for pos in self.positions.values())
        return max_investment - current_investment
    
    def can_buy(self, nlv):
        position_value = nlv * self.position_size_pct
        buying_power = self.get_buying_power(nlv)
        return buying_power >= position_value
    
    def buy(self, ticker, price, date, nlv):
        if ticker in self.positions:
            return False
        
        if ticker in self.pre_existing:
            return False
        
        position_value = nlv * self.position_size_pct
        buying_power = self.get_buying_power(nlv)
        
        if buying_power < position_value:
            return False
        
        shares = int(position_value / price)
        if shares < 1:
            return False
        
        actual_cost = shares * price
        
        if actual_cost > self.cash:
            margin_used = actual_cost - self.cash
            self.cash = 0
        else:
            margin_used = 0
            self.cash -= actual_cost
        
        self.positions[ticker] = {
            'shares': shares,
            'entry_price': price,
            'entry_date': date,
            'margin_used': margin_used
        }
        
        self.trade_history.append({
            'date': date,
            'ticker': ticker,
            'action': 'BUY',
            'shares': shares,
            'price': price,
            'value': actual_cost,
            'margin_used': margin_used,
            'nlv': nlv
        })
        
        return True
    
    def sell(self, ticker, price, date, nlv):
        if ticker in self.pre_existing:
            self.pre_existing.remove(ticker)
            return False
            
        if ticker not in self.positions:
            return False
        
        pos = self.positions[ticker]
        proceeds = pos['shares'] * price
        
        if pos['margin_used'] > 0:
            self.cash += proceeds - pos['margin_used']
        else:
            self.cash += proceeds
        
        # Calculate holding period
        holding_days = (date - pos['entry_date']).days
        
        self.trade_history.append({
            'date': date,
            'ticker': ticker,
            'action': 'SELL',
            'shares': pos['shares'],
            'price': price,
            'value': proceeds,
            'pnl': proceeds - (pos['shares'] * pos['entry_price']),
            'pnl_pct': ((price - pos['entry_price']) / pos['entry_price']) * 100,
            'holding_days': holding_days,
            'nlv': nlv
        })
        
        del self.positions[ticker]
        return True

# === Run simulation for a strategy ===
def run_simulation(ticker_data, portfolio_dates, signal_column, strategy_name):
    # Check for pre-existing positions
    pre_existing_positions = set()
    for ticker in ticker_data:
        df = ticker_data[ticker]
        position_open = False
        
        for date in df.index:
            if date >= pd.to_datetime(PORTFOLIO_START_DATE):
                break
            if df.loc[date, signal_column] == 1:
                position_open = True
            elif df.loc[date, signal_column] == -1:
                position_open = False
        
        if position_open:
            pre_existing_positions.add(ticker)
    
    # Initialize portfolio
    portfolio = Portfolio(INITIAL_CAPITAL, MAX_LEVERAGE, POSITION_SIZE_PCT, strategy_name)
    portfolio.pre_existing = pre_existing_positions.copy()
    
    # Run simulation
    for date in portfolio_dates:
        nlv = portfolio.get_nlv(date, ticker_data)
        
        # Store daily value
        portfolio.daily_values.append({'date': date, 'nlv': nlv})
        
        # Check for sell signals first
        for ticker in list(portfolio.positions.keys()):
            if ticker in ticker_data and date in ticker_data[ticker].index:
                if ticker_data[ticker].loc[date, signal_column] == -1:
                    price = ticker_data[ticker].loc[date, "Close"]
                    portfolio.sell(ticker, price, date, nlv)
        
        # Check pre-existing positions for sells
        for ticker in list(portfolio.pre_existing):
            if ticker in ticker_data and date in ticker_data[ticker].index:
                if ticker_data[ticker].loc[date, signal_column] == -1:
                    portfolio.pre_existing.remove(ticker)
        
        # Update NLV after sells
        nlv = portfolio.get_nlv(date, ticker_data)
        
        # Check for buy signals
        buy_candidates = []
        for ticker in ticker_data:
            if ticker not in portfolio.positions and ticker not in portfolio.pre_existing:
                if date in ticker_data[ticker].index:
                    if ticker_data[ticker].loc[date, signal_column] == 1:
                        price = ticker_data[ticker].loc[date, "Close"]
                        buy_candidates.append((ticker, price))
        
        # Randomize and try to buy
        if buy_candidates:
            random.shuffle(buy_candidates)
            for ticker, price in buy_candidates:
                if portfolio.can_buy(nlv):
                    portfolio.buy(ticker, price, date, nlv)
                    nlv = portfolio.get_nlv(date, ticker_data)
    
    return portfolio

# === Calculate performance metrics ===
def calculate_metrics(portfolio, final_date, ticker_data):
    final_nlv = portfolio.get_nlv(final_date, ticker_data)
    
    # Basic metrics
    total_return = final_nlv - INITIAL_CAPITAL
    return_pct = (total_return / INITIAL_CAPITAL) * 100
    
    # Trade metrics
    completed_trades = [t for t in portfolio.trade_history if t['action'] == 'SELL']
    if completed_trades:
        winning_trades = [t for t in completed_trades if t['pnl'] > 0]
        losing_trades = [t for t in completed_trades if t['pnl'] <= 0]
        win_rate = (len(winning_trades) / len(completed_trades)) * 100
        
        avg_win = np.mean([t['pnl'] for t in winning_trades]) if winning_trades else 0
        avg_loss = np.mean([t['pnl'] for t in losing_trades]) if losing_trades else 0
        avg_holding = np.mean([t['holding_days'] for t in completed_trades])
        
        # Profit factor
        total_wins = sum(t['pnl'] for t in winning_trades) if winning_trades else 0
        total_losses = abs(sum(t['pnl'] for t in losing_trades)) if losing_trades else 1
        profit_factor = total_wins / total_losses if total_losses > 0 else 0
    else:
        win_rate = avg_win = avg_loss = avg_holding = profit_factor = 0
        winning_trades = []
        losing_trades = []
    
    # Maximum drawdown
    daily_df = pd.DataFrame(portfolio.daily_values)
    if not daily_df.empty:
        daily_df['peak'] = daily_df['nlv'].cummax()
        daily_df['drawdown'] = (daily_df['nlv'] - daily_df['peak']) / daily_df['peak']
        max_drawdown = daily_df['drawdown'].min() * 100
    else:
        max_drawdown = 0
    
    return {
        'final_nlv': final_nlv,
        'total_return': total_return,
        'return_pct': return_pct,
        'total_trades': len(portfolio.trade_history),
        'completed_trades': len(completed_trades),
        'open_positions': len(portfolio.positions),
        'winning_trades': len(winning_trades),
        'losing_trades': len(losing_trades),
        'win_rate': win_rate,
        'avg_win': avg_win,
        'avg_loss': avg_loss,
        'profit_factor': profit_factor,
        'avg_holding_days': avg_holding,
        'max_drawdown': max_drawdown
    }

# === MAIN EXECUTION ===

# Load data
print("\n📂 Loading ticker data...")
ticker_data, all_dates = load_ticker_data(tickers, DATA_START_DATE, END_DATE)
portfolio_dates = [d for d in all_dates if d >= pd.to_datetime(PORTFOLIO_START_DATE)]
final_date = portfolio_dates[-1] if portfolio_dates else all_dates[-1]

print(f"✅ Loaded {len(ticker_data)} tickers successfully")

# Apply both strategies
print("\n🔄 Applying trading strategies...")
apply_original_strategy(ticker_data)
apply_two_closes_strategy(ticker_data)

# Run simulations
print("\n📈 Running Strategy 1: Original (Two-step exit)...")
portfolio_original = run_simulation(ticker_data, portfolio_dates, "signal_original", "Original")

print("\n📈 Running Strategy 2: Two Closes Below EMA...")
portfolio_2closes = run_simulation(ticker_data, portfolio_dates, "signal_2closes", "2 Closes Below")

# Calculate metrics
metrics_original = calculate_metrics(portfolio_original, final_date, ticker_data)
metrics_2closes = calculate_metrics(portfolio_2closes, final_date, ticker_data)

# === COMPARISON REPORT ===
print("\n" + "="*80)
print("📊 STRATEGY COMPARISON REPORT")
print("="*80)

# Create comparison table
comparison_data = {
    'Metric': [
        'Final Portfolio Value',
        'Total Return ($)',
        'Total Return (%)',
        'Total Trades',
        'Completed Trades',
        'Open Positions',
        'Winning Trades',
        'Losing Trades',
        'Win Rate (%)',
        'Average Win ($)',
        'Average Loss ($)',
        'Profit Factor',
        'Avg Holding Days',
        'Max Drawdown (%)'
    ],
    'Original Strategy': [
        f"${metrics_original['final_nlv']:,.2f}",
        f"${metrics_original['total_return']:,.2f}",
        f"{metrics_original['return_pct']:.2f}%",
        metrics_original['total_trades'],
        metrics_original['completed_trades'],
        metrics_original['open_positions'],
        metrics_original['winning_trades'],
        metrics_original['losing_trades'],
        f"{metrics_original['win_rate']:.2f}%",
        f"${metrics_original['avg_win']:,.2f}",
        f"${metrics_original['avg_loss']:,.2f}",
        f"{metrics_original['profit_factor']:.2f}",
        f"{metrics_original['avg_holding_days']:.1f}",
        f"{metrics_original['max_drawdown']:.2f}%"
    ],
    '2 Closes Below EMA': [
        f"${metrics_2closes['final_nlv']:,.2f}",
        f"${metrics_2closes['total_return']:,.2f}",
        f"{metrics_2closes['return_pct']:.2f}%",
        metrics_2closes['total_trades'],
        metrics_2closes['completed_trades'],
        metrics_2closes['open_positions'],
        metrics_2closes['winning_trades'],
        metrics_2closes['losing_trades'],
        f"{metrics_2closes['win_rate']:.2f}%",
        f"${metrics_2closes['avg_win']:,.2f}",
        f"${metrics_2closes['avg_loss']:,.2f}",
        f"{metrics_2closes['profit_factor']:.2f}",
        f"{metrics_2closes['avg_holding_days']:.1f}",
        f"{metrics_2closes['max_drawdown']:.2f}%"
    ]
}

comparison_df = pd.DataFrame(comparison_data)

# Display comparison
print("\n" + comparison_df.to_string(index=False))

# Determine winner
print("\n" + "="*80)
print("📊 ANALYSIS SUMMARY")
print("="*80)

better_return = "Original" if metrics_original['return_pct'] > metrics_2closes['return_pct'] else "2 Closes Below"
better_winrate = "Original" if metrics_original['win_rate'] > metrics_2closes['win_rate'] else "2 Closes Below"
better_drawdown = "Original" if metrics_original['max_drawdown'] > metrics_2closes['max_drawdown'] else "2 Closes Below"

print(f"\n✅ Better Total Return: {better_return} Strategy")
print(f"   Original: {metrics_original['return_pct']:.2f}% vs 2 Closes: {metrics_2closes['return_pct']:.2f}%")
print(f"   Difference: {abs(metrics_original['return_pct'] - metrics_2closes['return_pct']):.2f}%")

print(f"\n✅ Better Win Rate: {better_winrate} Strategy")
print(f"   Original: {metrics_original['win_rate']:.2f}% vs 2 Closes: {metrics_2closes['win_rate']:.2f}%")

print(f"\n✅ Lower Drawdown: {better_drawdown} Strategy")
print(f"   Original: {metrics_original['max_drawdown']:.2f}% vs 2 Closes: {metrics_2closes['max_drawdown']:.2f}%")

# Trade frequency analysis
print(f"\n📊 Trade Frequency:")
print(f"   Original: {metrics_original['completed_trades']} completed trades")
print(f"   2 Closes: {metrics_2closes['completed_trades']} completed trades")
print(f"   Difference: {abs(metrics_original['completed_trades'] - metrics_2closes['completed_trades'])} trades")

# Save comparison to Excel
print("\n💾 Saving detailed comparison...")
with pd.ExcelWriter("output/strategy_comparison.xlsx", engine='openpyxl') as writer:
    # Summary sheet
    comparison_df.to_excel(writer, sheet_name='Comparison', index=False)
    
    # Original strategy trades
    if portfolio_original.trade_history:
        original_trades_df = pd.DataFrame(portfolio_original.trade_history)
        original_trades_df.to_excel(writer, sheet_name='Original Trades', index=False)
    
    # 2 Closes strategy trades
    if portfolio_2closes.trade_history:
        closes_trades_df = pd.DataFrame(portfolio_2closes.trade_history)
        closes_trades_df.to_excel(writer, sheet_name='2 Closes Trades', index=False)
    
    # Format the comparison sheet
    workbook = writer.book
    worksheet = writer.sheets['Comparison']
    
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Format headers
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Highlight better values
    for row in range(2, len(comparison_data['Metric']) + 2):
        metric_cell = worksheet.cell(row=row, column=1)
        original_cell = worksheet.cell(row=row, column=2)
        closes_cell = worksheet.cell(row=row, column=3)
        
        # Color code based on which is better
        if metric_cell.value in ['Total Return ($)', 'Total Return (%)', 'Win Rate (%)', 'Profit Factor']:
            # Higher is better
            if '$' in str(original_cell.value):
                orig_val = float(str(original_cell.value).replace('$', '').replace(',', '').replace('%', ''))
                close_val = float(str(closes_cell.value).replace('$', '').replace(',', '').replace('%', ''))
            else:
                try:
                    orig_val = float(str(original_cell.value).replace('%', ''))
                    close_val = float(str(closes_cell.value).replace('%', ''))
                except:
                    continue
            
            if orig_val > close_val:
                original_cell.font = Font(color="008000", bold=True)
            elif close_val > orig_val:
                closes_cell.font = Font(color="008000", bold=True)
        
        elif metric_cell.value == 'Max Drawdown (%)':
            # Lower is better (less negative)
            try:
                orig_val = float(str(original_cell.value).replace('%', ''))
                close_val = float(str(closes_cell.value).replace('%', ''))
                if orig_val > close_val:  # Less negative is better
                    original_cell.font = Font(color="008000", bold=True)
                elif close_val > orig_val:
                    closes_cell.font = Font(color="008000", bold=True)
            except:
                continue

print("✅ Comparison saved to: output/strategy_comparison.xlsx")

# === Create ticker-by-ticker comparison ===
print("\n📊 Analyzing performance by ticker...")

ticker_performance = []
for ticker in ticker_data:
    df = ticker_data[ticker]
    
    # Count signals for each strategy
    original_buys = len(df[df['signal_original'] == 1])
    original_sells = len(df[df['signal_original'] == -1])
    closes_buys = len(df[df['signal_2closes'] == 1])
    closes_sells = len(df[df['signal_2closes'] == -1])
    
    ticker_performance.append({
        'Ticker': ticker,
        'Original_Buys': original_buys,
        'Original_Sells': original_sells,
        '2Closes_Buys': closes_buys,
        '2Closes_Sells': closes_sells,
        'Buy_Diff': closes_buys - original_buys,
        'Sell_Diff': closes_sells - original_sells
    })

ticker_perf_df = pd.DataFrame(ticker_performance)
print("\nSignal Count by Ticker:")
print(ticker_perf_df.to_string(index=False))

# Save ticker analysis
ticker_perf_df.to_csv("output/ticker_signal_comparison.csv", index=False)
print("\n💾 Ticker signal analysis saved to: output/ticker_signal_comparison.csv")

print("\n✅ Backtest comparison complete!")
print("📁 All results saved to the 'output' folder")
print("\n📌 KEY FINDINGS:")
print(f"   • The {better_return} strategy produced better returns")
print(f"   • The {better_winrate} strategy had a higher win rate")
print(f"   • The {better_drawdown} strategy had lower drawdown risk")
print(f"\n💡 Recommendation: Consider your risk tolerance and trading goals when choosing between strategies.")