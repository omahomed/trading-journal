import pandas as pd
import numpy as np
import warnings
import os
from datetime import datetime
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === Setup ===
warnings.simplefilter(action="ignore", category=FutureWarning)

# === User Configuration ===
INITIAL_CAPITAL = 150000
MAX_LEVERAGE = 1.8
POSITION_SIZE_PCT = 0.10
START_DATE = "2024-09-04"
END_DATE = "2025-09-05"

# === Load tickers ===
with open("select_ticker.txt", "r") as f:
    tickers = [line.strip().upper() for line in f if line.strip()]

print("📊 Weekly Summary Report Generator")
print(f"💰 Initial Capital: ${INITIAL_CAPITAL:,.2f}")
print(f"📈 Max Leverage: {MAX_LEVERAGE:.1%}")
print(f"📅 Period: {START_DATE} to {END_DATE}\n")

# === Load ticker data ===
ticker_data = {}
all_dates = set()

for ticker in tickers:
    input_file = f"output/{ticker}_price_data.csv"
    
    try:
        df = pd.read_csv(input_file)
        df["Date"] = pd.to_datetime(df["Date"])
        df.set_index("Date", inplace=True)
        df = df[(df.index >= START_DATE) & (df.index <= END_DATE)]
        
        if df.empty:
            continue
        
        for col in ["Open", "High", "Low", "Close"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        df.dropna(subset=["Open", "High", "Low", "Close"], inplace=True)
        
        # Calculate indicators
        df["EMA21"] = df["Close"].ewm(span=21, adjust=False).mean()
        df["sell_threshold"] = df["EMA21"] * 0.998
        df["low_above_ema"] = df["Low"] > df["EMA21"]
        df["up_day"] = df["Close"] > df["Close"].shift(1)
        
        # Generate signals
        df["signal"] = 0
        streak = 0
        in_wait_mode = False
        position_open = False
        
        for i in range(2, len(df)):
            if position_open:
                if df.iloc[i]["Close"] < df.iloc[i]["sell_threshold"]:
                    df.at[df.index[i], "signal"] = -1
                    position_open = False
                    in_wait_mode = False
                    streak = 0
                    continue
            
            if df.iloc[i]["low_above_ema"]:
                streak += 1
            elif df.iloc[i]["Close"] < df.iloc[i]["EMA21"]:
                streak = 0
                in_wait_mode = False
            
            if streak >= 3 and not position_open:
                if df.iloc[i]["up_day"]:
                    df.at[df.index[i], "signal"] = 1
                    position_open = True
                    streak = 0
                    in_wait_mode = False
                else:
                    in_wait_mode = True
            
            elif in_wait_mode and df.iloc[i]["up_day"] and df.iloc[i]["low_above_ema"] and not position_open:
                df.at[df.index[i], "signal"] = 1
                position_open = True
                streak = 0
                in_wait_mode = False
        
        ticker_data[ticker] = df
        all_dates.update(df.index)
        print(f"✅ Loaded {ticker}: {len(df)} days")
    except:
        continue

all_dates = sorted(list(all_dates))
print(f"\n📅 Total trading days: {len(all_dates)}")

# === Simulate with entry date tracking ===
print("🔄 Running simulation with entry date tracking...\n")

positions = {}  # {ticker: {'shares', 'entry_price', 'entry_date', 'cost_basis'}}
cash = INITIAL_CAPITAL
trades = []
daily_snapshots = []

for date in all_dates:
    # Process sells
    for ticker in list(positions.keys()):
        if ticker in ticker_data and date in ticker_data[ticker].index:
            if ticker_data[ticker].loc[date, "signal"] == -1:
                pos = positions[ticker]
                exit_price = ticker_data[ticker].loc[date, "Close"]
                proceeds = pos['shares'] * exit_price
                cash += proceeds
                
                pnl = proceeds - pos['cost_basis']
                pnl_pct = (pnl / pos['cost_basis'] * 100)
                
                trades.append({
                    'date': date,
                    'ticker': ticker,
                    'action': 'SELL',
                    'shares': pos['shares'],
                    'exit_price': exit_price,
                    'entry_date': pos['entry_date'],
                    'entry_price': pos['entry_price'],
                    'pnl': pnl,
                    'pnl_pct': pnl_pct
                })
                
                print(f"📉 SELL {ticker} on {date.date()} (entered {pos['entry_date'].date()}) - P&L: ${pnl:.2f}")
                del positions[ticker]
    
    # Process buys
    buy_candidates = []
    for ticker in ticker_data:
        if ticker not in positions:
            if date in ticker_data[ticker].index:
                if ticker_data[ticker].loc[date, "signal"] == 1:
                    buy_candidates.append((ticker, ticker_data[ticker].loc[date, "Close"]))
    
    if buy_candidates:
        random.shuffle(buy_candidates)
        
        for ticker, price in buy_candidates:
            nlv = cash
            for t, p in positions.items():
                if t in ticker_data and date in ticker_data[t].index:
                    nlv += p['shares'] * ticker_data[t].loc[date, "Close"]
            
            position_size = nlv * POSITION_SIZE_PCT
            current_investment = sum(p['cost_basis'] for p in positions.values())
            buying_power = (nlv * MAX_LEVERAGE) - current_investment
            
            if buying_power >= position_size:
                shares = int(position_size / price)
                if shares > 0:
                    cost = shares * price
                    cash -= min(cost, cash)
                    
                    positions[ticker] = {
                        'shares': shares,
                        'entry_price': price,
                        'entry_date': date,  # STORE THE DATE!
                        'cost_basis': cost
                    }
                    
                    trades.append({
                        'date': date,
                        'ticker': ticker,
                        'action': 'BUY',
                        'shares': shares,
                        'price': price
                    })
                    
                    print(f"📈 BUY {ticker} on {date.date()}")
    
    # Daily snapshot
    nlv = cash
    position_details = {}
    
    for ticker, pos in positions.items():
        if ticker in ticker_data and date in ticker_data[ticker].index:
            current_price = ticker_data[ticker].loc[date, "Close"]
            current_value = pos['shares'] * current_price
            nlv += current_value
            
            pnl = current_value - pos['cost_basis']
            pnl_pct = (pnl / pos['cost_basis'] * 100)
            
            position_details[ticker] = {
                'shares': pos['shares'],
                'entry_price': pos['entry_price'],
                'entry_date': pos['entry_date'],
                'current_price': current_price,
                'cost_basis': pos['cost_basis'],
                'current_value': current_value,
                'pnl': pnl,
                'pnl_pct': pnl_pct
            }
    
    daily_snapshots.append({
        'date': date,
        'week': f"{date.year}-W{date.isocalendar()[1]:02d}",
        'nlv': nlv,
        'cash': cash,
        'positions': position_details.copy()
    })

# === Create Weekly Summary ===
print(f"\n✅ Simulation complete. Creating weekly summary...\n")

snapshots_df = pd.DataFrame(daily_snapshots)
weekly_summaries = []

# Group by week and sort chronologically
week_groups = []
for week, week_data in snapshots_df.groupby('week'):
    week_data = week_data.sort_values('date')
    first_date = week_data.iloc[0]['date']
    week_groups.append((first_date, week, week_data))

# Sort by actual date, not week string
week_groups.sort(key=lambda x: x[0])

for first_date, week, week_data in week_groups:
    week_data = week_data.sort_values('date')
    first_day = week_data.iloc[0]
    last_day = week_data.iloc[-1]
    
    week_start = first_day['date']
    week_end = last_day['date']
    week_trades = [t for t in trades if week_start <= t['date'] <= week_end]
    
    # Calculate cumulative P&L
    cumulative_pnl = sum(pos['pnl'] for pos in last_day['positions'].values()) if last_day['positions'] else 0
    
    summary = {
        'week': week,
        'start_date': week_start,
        'end_date': week_end,
        'beginning_nlv': first_day['nlv'],
        'ending_nlv': last_day['nlv'],
        'week_change': last_day['nlv'] - first_day['nlv'],
        'week_change_pct': ((last_day['nlv'] - first_day['nlv']) / first_day['nlv'] * 100) if first_day['nlv'] > 0 else 0,
        'cash': last_day['cash'],
        'cumulative_open_pnl': cumulative_pnl,
        'num_trades': len(week_trades),
        'positions': last_day['positions'],
        'closed_trades': [t for t in week_trades if t['action'] == 'SELL']
    }
    
    weekly_summaries.append(summary)

# === Create Excel Report ===
wb = Workbook()
ws = wb.active
ws.title = "Weekly Summary"

# Styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
label_font = Font(bold=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), 
               top=Side(style='thin'), bottom=Side(style='thin'))

row = 1
ws.cell(row=row, column=1, value="WEEKLY PORTFOLIO SUMMARY").font = Font(bold=True, size=14)
row += 2

# === Create Excel Report ===
wb = Workbook()
ws = wb.active
ws.title = "Weekly Summary"

# Styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
label_font = Font(bold=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), 
               top=Side(style='thin'), bottom=Side(style='thin'))

row = 1
ws.cell(row=row, column=1, value="WEEKLY PORTFOLIO SUMMARY").font = Font(bold=True, size=14)
row += 2

# Show all weeks (or limit as needed)
for week_num, week in enumerate(weekly_summaries, 1):
    # Week header - use sequential week number instead of ISO week
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws.cell(row=row, column=1, 
                  value=f"WEEK {week_num} ({week['start_date'].strftime('%Y-%m-%d')} to {week['end_date'].strftime('%Y-%m-%d')})")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Metrics
    ws.cell(row=row, column=1, value="Beginning Balance:").font = label_font
    ws.cell(row=row, column=3, value=f"${week['beginning_nlv']:,.2f}")
    row += 1
    
    ws.cell(row=row, column=1, value="Ending Balance:").font = label_font
    ws.cell(row=row, column=3, value=f"${week['ending_nlv']:,.2f}")
    row += 1
    
    ws.cell(row=row, column=1, value="Week Change:").font = label_font
    ws.cell(row=row, column=3, value=f"${week['week_change']:,.2f} ({week['week_change_pct']:.2f}%)")
    row += 1
    
    ws.cell(row=row, column=1, value="Cumulative Open P&L:").font = label_font
    ws.cell(row=row, column=3, value=f"${week['cumulative_open_pnl']:,.2f}")
    row += 2
    
    # Active Positions
    if week['positions']:
        ws.cell(row=row, column=1, value="ACTIVE POSITIONS:").font = label_font
        row += 1
        
        headers = ["Ticker", "Entry Date", "Shares", "Entry Price", "Current Price", "Cost Basis", "Current Value", "P&L $", "P&L %"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).font = label_font
        row += 1
        
        for ticker, pos in week['positions'].items():
            ws.cell(row=row, column=1, value=ticker)
            ws.cell(row=row, column=2, value=pos['entry_date'].strftime('%Y-%m-%d'))
            ws.cell(row=row, column=3, value=pos['shares'])
            ws.cell(row=row, column=4, value=f"${pos['entry_price']:.2f}")
            ws.cell(row=row, column=5, value=f"${pos['current_price']:.2f}")
            ws.cell(row=row, column=6, value=f"${pos['cost_basis']:,.2f}")
            ws.cell(row=row, column=7, value=f"${pos['current_value']:,.2f}")
            ws.cell(row=row, column=8, value=f"${pos['pnl']:,.2f}")
            ws.cell(row=row, column=9, value=f"{pos['pnl_pct']:.2f}%")
            row += 1
    
    row += 1
    
    # Closed Trades
    if week['closed_trades']:
        ws.cell(row=row, column=1, value="CLOSED TRADES:").font = label_font
        row += 1
        
        headers = ["Ticker", "Entry Date", "Exit Date", "Shares", "Entry Price", "Exit Price", "P&L $", "P&L %"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).font = label_font
        row += 1
        
        for trade in week['closed_trades']:
            ws.cell(row=row, column=1, value=trade['ticker'])
            ws.cell(row=row, column=2, value=trade['entry_date'].strftime('%Y-%m-%d'))
            ws.cell(row=row, column=3, value=trade['date'].strftime('%Y-%m-%d'))
            ws.cell(row=row, column=4, value=trade['shares'])
            ws.cell(row=row, column=5, value=f"${trade['entry_price']:.2f}")
            ws.cell(row=row, column=6, value=f"${trade['exit_price']:.2f}")
            ws.cell(row=row, column=7, value=f"${trade['pnl']:,.2f}")
            ws.cell(row=row, column=8, value=f"{trade['pnl_pct']:.2f}%")
            row += 1
    
    row += 3  # Space between weeks

# Auto-adjust columns
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

# Save
wb.save("output/weekly_summary_with_dates.xlsx")
print("💾 Saved to: output/weekly_summary_with_dates.xlsx")
print(f"📊 Total weeks: {len(weekly_summaries)}")
print("\n✅ Done! Check the Excel file for entry dates and closed trades.")