import pandas as pd
import numpy as np
import warnings
import os
from datetime import datetime
import random

# === Setup ===
warnings.simplefilter(action="ignore", category=FutureWarning)
os.makedirs("output", exist_ok=True)

# === User Configuration ===
INITIAL_CAPITAL = 150000  # Starting portfolio value
MAX_LEVERAGE = 1.8  # Can invest up to 180% of portfolio value
POSITION_SIZE_PCT = 0.10  # 10% of NLV per position
DATA_START_DATE = "2023-09-01"  # Load data from here (for signal calculation)
PORTFOLIO_START_DATE = "2024-09-04"  # Start portfolio trading from here
END_DATE = "2025-09-05"

# === Load tickers ===
with open("select_ticker.txt", "r") as f:
    tickers = [line.strip().upper() for line in f if line.strip()]

print(f"📊 Portfolio Simulation with Margin")
print(f"💰 Initial Capital: ${INITIAL_CAPITAL:,.2f}")
print(f"📈 Max Leverage: {MAX_LEVERAGE:.1%}")
print(f"📦 Position Size: {POSITION_SIZE_PCT:.1%} of NLV")
print(f"📅 Data Period: {DATA_START_DATE} to {END_DATE}")
print(f"💼 Portfolio Start: {PORTFOLIO_START_DATE}")
print(f"🎯 Tickers: {len(tickers)} available\n")

# === Load and prepare all ticker data ===
ticker_data = {}
all_dates = set()

for ticker in tickers:
    input_file = f"output/{ticker}_price_data.csv"
    
    try:
        df = pd.read_csv(input_file)
    except FileNotFoundError:
        print(f"⚠️ File not found: {input_file}. Skipping {ticker}.")
        continue
    
    # Filter date range - load from DATA_START_DATE for signal calculation
    df["Date"] = pd.to_datetime(df["Date"])
    df.set_index("Date", inplace=True)
    df = df[(df.index >= DATA_START_DATE) & (df.index <= END_DATE)]
    
    if df.empty:
        print(f"⚠️ No data in date range for {ticker}. Skipping.")
        continue
    
    # Clean data
    for col in ["Open", "High", "Low", "Close"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df.dropna(subset=["Open", "High", "Low", "Close"], inplace=True)
    
    # Calculate indicators
    df["EMA21"] = df["Close"].ewm(span=21, adjust=False).mean()
    
    # Generate signals using Pine Script logic
    df["signal"] = 0
    position_open = False
    first_break_low = None
    
    for i in range(2, len(df)):
        # BUY SIGNAL: 3 consecutive days with low > EMA21 AND current day is up day (close > previous close)
        if not position_open and i >= 2:
            three_days_above = (df.iloc[i]["Low"] > df.iloc[i]["EMA21"] and 
                               df.iloc[i-1]["Low"] > df.iloc[i-1]["EMA21"] and 
                               df.iloc[i-2]["Low"] > df.iloc[i-2]["EMA21"])
            is_up_day = df.iloc[i]["Close"] > df.iloc[i-1]["Close"]  # FIXED: Compare to previous close, not open
            
            if three_days_above and is_up_day:
                df.at[df.index[i], "signal"] = 1
                position_open = True
                first_break_low = None
        
        # SELL SIGNAL: Two-step process
        elif position_open:
            # Step 1: Check if we closed below EMA for first time
            if first_break_low is None:
                if df.iloc[i]["Close"] < df.iloc[i]["EMA21"]:
                    if i > 0 and df.iloc[i-1]["Close"] >= df.iloc[i-1]["EMA21"]:
                        first_break_low = df.iloc[i]["Low"]
            
            # Step 2: Sell if low drops below the stored low
            if first_break_low is not None:
                if df.iloc[i]["Low"] < first_break_low:
                    df.at[df.index[i], "signal"] = -1
                    position_open = False
                    first_break_low = None
    
    ticker_data[ticker] = df
    all_dates.update(df.index)

# Sort dates and filter for portfolio period
all_dates = sorted(list(all_dates))
portfolio_dates = [d for d in all_dates if d >= pd.to_datetime(PORTFOLIO_START_DATE)]

# Check for pre-existing positions
print("🔍 Checking for pre-existing positions...")
pre_existing_positions = set()
for ticker in ticker_data:
    df = ticker_data[ticker]
    position_open = False
    
    for date in df.index:
        if date >= pd.to_datetime(PORTFOLIO_START_DATE):
            break
        if df.loc[date, "signal"] == 1:
            position_open = True
        elif df.loc[date, "signal"] == -1:
            position_open = False
    
    if position_open:
        pre_existing_positions.add(ticker)

if pre_existing_positions:
    print(f"⚠️ Found {len(pre_existing_positions)} tickers already in position at portfolio start:")
    for ticker in pre_existing_positions:
        print(f"   - {ticker} (will wait for sell signal before allowing new buy)")
else:
    print("✅ No pre-existing positions found")

print()

# === Portfolio Simulation ===
class Portfolio:
    def __init__(self, initial_capital, max_leverage, position_size_pct):
        self.initial_capital = initial_capital
        self.cash = initial_capital
        self.max_leverage = max_leverage
        self.position_size_pct = position_size_pct
        self.positions = {}  # {ticker: {'shares': x, 'entry_price': y, 'entry_date': z}}
        self.trade_history = []
        self.pre_existing = pre_existing_positions.copy()
        
    def get_nlv(self, date, ticker_data):
        """Calculate Net Liquidating Value (cash + value of all positions)"""
        nlv = self.cash
        for ticker, pos in self.positions.items():
            if ticker in ticker_data and date in ticker_data[ticker].index:
                current_price = ticker_data[ticker].loc[date, "Close"]
                nlv += pos['shares'] * current_price
        return nlv
    
    def get_buying_power(self, nlv):
        """Calculate available buying power based on margin"""
        max_investment = nlv * self.max_leverage
        current_investment = sum(pos['shares'] * pos['entry_price'] for pos in self.positions.values())
        return max_investment - current_investment
    
    def can_buy(self, nlv):
        """Check if we have room for a new position"""
        position_value = nlv * self.position_size_pct
        buying_power = self.get_buying_power(nlv)
        return buying_power >= position_value
    
    def buy(self, ticker, price, date, nlv):
        """Execute a buy order"""
        if ticker in self.positions:
            return False  # Already have position
        
        if ticker in self.pre_existing:
            return False  # Skip pre-existing positions until they clear
        
        position_value = nlv * self.position_size_pct
        buying_power = self.get_buying_power(nlv)
        
        if buying_power < position_value:
            return False  # Not enough buying power
        
        shares = int(position_value / price)
        if shares < 1:
            return False
        
        actual_cost = shares * price
        
        # Use margin if needed
        if actual_cost > self.cash:
            margin_used = actual_cost - self.cash
            self.cash = 0
        else:
            margin_used = 0
            self.cash -= actual_cost
        
        self.positions[ticker] = {
            'shares': shares,
            'entry_price': price,
            'entry_date': date,
            'margin_used': margin_used
        }
        
        self.trade_history.append({
            'date': date,
            'ticker': ticker,
            'action': 'BUY',
            'shares': shares,
            'price': price,
            'value': actual_cost,
            'margin_used': margin_used,
            'nlv': nlv
        })
        
        return True
    
    def sell(self, ticker, price, date, nlv):
        """Execute a sell order"""
        # Remove from pre-existing if it was there
        if ticker in self.pre_existing:
            self.pre_existing.remove(ticker)
            return False  # Don't track the sale of pre-existing positions
            
        if ticker not in self.positions:
            return False
        
        pos = self.positions[ticker]
        proceeds = pos['shares'] * price
        
        # Pay back margin if used
        if pos['margin_used'] > 0:
            self.cash += proceeds - pos['margin_used']
        else:
            self.cash += proceeds
        
        self.trade_history.append({
            'date': date,
            'ticker': ticker,
            'action': 'SELL',
            'shares': pos['shares'],
            'price': price,
            'value': proceeds,
            'pnl': proceeds - (pos['shares'] * pos['entry_price']),
            'nlv': nlv
        })
        
        del self.positions[ticker]
        return True

# Initialize portfolio
portfolio = Portfolio(INITIAL_CAPITAL, MAX_LEVERAGE, POSITION_SIZE_PCT)

# === Run simulation day by day (ONLY FROM PORTFOLIO START DATE) ===
print("🔄 Running simulation...\n")

for date in portfolio_dates:  # Use portfolio_dates, not all_dates!
    # Get current NLV
    nlv = portfolio.get_nlv(date, ticker_data)
    
    # Check for sell signals first (free up capital)
    for ticker in list(portfolio.positions.keys()):
        if ticker in ticker_data and date in ticker_data[ticker].index:
            if ticker_data[ticker].loc[date, "signal"] == -1:
                price = ticker_data[ticker].loc[date, "Close"]
                if portfolio.sell(ticker, price, date, nlv):
                    print(f"📉 {date.date()} SELL {ticker} @ ${price:.2f}")
    
    # Also check pre-existing positions for sells
    for ticker in list(portfolio.pre_existing):
        if ticker in ticker_data and date in ticker_data[ticker].index:
            if ticker_data[ticker].loc[date, "signal"] == -1:
                portfolio.pre_existing.remove(ticker)
                print(f"📉 {date.date()} {ticker} pre-existing position cleared")
    
    # Update NLV after sells
    nlv = portfolio.get_nlv(date, ticker_data)
    
    # Check for buy signals
    buy_candidates = []
    for ticker in ticker_data:
        if ticker not in portfolio.positions and ticker not in portfolio.pre_existing:
            if date in ticker_data[ticker].index:
                if ticker_data[ticker].loc[date, "signal"] == 1:
                    price = ticker_data[ticker].loc[date, "Close"]
                    buy_candidates.append((ticker, price))
    
    # Randomize order of buy candidates
    if buy_candidates:
        random.shuffle(buy_candidates)
        
        # Try to buy as many as we can afford
        for ticker, price in buy_candidates:
            if portfolio.can_buy(nlv):
                if portfolio.buy(ticker, price, date, nlv):
                    print(f"📈 {date.date()} BUY  {ticker} @ ${price:.2f}")
                    # Update NLV for next potential buy
                    nlv = portfolio.get_nlv(date, ticker_data)

# === Calculate final results ===
final_date = portfolio_dates[-1] if portfolio_dates else all_dates[-1]
final_nlv = portfolio.get_nlv(final_date, ticker_data)

print("\n" + "="*60)
print("📊 FINAL RESULTS")
print("="*60)
print(f"Initial Capital: ${INITIAL_CAPITAL:,.2f}")
print(f"Final NLV: ${final_nlv:,.2f}")
print(f"Total Return: ${final_nlv - INITIAL_CAPITAL:,.2f}")
print(f"Return %: {((final_nlv - INITIAL_CAPITAL) / INITIAL_CAPITAL) * 100:.2f}%")
print(f"Total Trades: {len(portfolio.trade_history)}")

# Current positions
if portfolio.positions:
    print(f"\n📦 Open Positions ({len(portfolio.positions)}):")
    for ticker, pos in portfolio.positions.items():
        current_price = ticker_data[ticker].iloc[-1]["Close"]
        pnl = (current_price - pos['entry_price']) * pos['shares']
        pnl_pct = ((current_price - pos['entry_price']) / pos['entry_price']) * 100
        print(f"  {ticker}: {pos['shares']} shares @ ${pos['entry_price']:.2f} | "
              f"Current: ${current_price:.2f} | P&L: ${pnl:,.2f} ({pnl_pct:+.2f}%)")
else:
    print("\n✅ All positions closed")

# === Save detailed results ===
# Trade history
trades_df = pd.DataFrame(portfolio.trade_history)
if not trades_df.empty:
    trades_df.to_csv("output/margin_trade_history.csv", index=False)
    print(f"\n💾 Trade history saved to: output/margin_trade_history.csv")

# Create Trade Log Summary (round-trip trades) - FIXED VERSION
print("\n📊 Creating Trade Log Summary...")
trade_log = []
trade_num = 0
open_trades = {}  # Track open positions for matching

for trade in portfolio.trade_history:
    if trade['action'] == 'BUY':
        # Store the buy trade info
        open_trades[trade['ticker']] = {
            'entry_date': trade['date'],
            'entry_price': trade['price'],
            'shares': trade['shares'],
            'position_size': trade['value'],
            'nlv_at_entry': trade['nlv']
        }
    elif trade['action'] == 'SELL':
        # Match with the buy trade
        if trade['ticker'] in open_trades:
            buy_info = open_trades[trade['ticker']]
            trade_num += 1
            
            # Calculate gain/loss
            entry_value = buy_info['shares'] * buy_info['entry_price']
            exit_value = trade['value']
            gain_loss = exit_value - entry_value
            gain_loss_pct = (gain_loss / entry_value) * 100 if entry_value > 0 else 0
            
            trade_log.append({
                'Trade_Num': trade_num,
                'Ticker': trade['ticker'],
                'Entry_Date': buy_info['entry_date'].strftime('%Y-%m-%d'),
                'NLV_at_Entry': round(buy_info['nlv_at_entry'], 2),
                'Position_Size': round(buy_info['position_size'], 2),
                'Exit_Date': trade['date'].strftime('%Y-%m-%d'),
                'Gain_Loss_Dollars': round(gain_loss, 2),
                'Gain_Loss_Pct': round(gain_loss_pct, 2),
                'Status': 'Closed'
            })
            
            # Remove from open trades
            del open_trades[trade['ticker']]

# Add hypothetically closed positions at end date
for ticker, buy_info in open_trades.items():
    if ticker in ticker_data and final_date in ticker_data[ticker].index:
        trade_num += 1
        
        # Get the closing price on the final date
        exit_price = ticker_data[ticker].loc[final_date, "Close"]
        
        # Calculate gain/loss
        entry_value = buy_info['shares'] * buy_info['entry_price']
        exit_value = buy_info['shares'] * exit_price
        gain_loss = exit_value - entry_value
        gain_loss_pct = (gain_loss / entry_value) if entry_value > 0 else 0
        
        trade_log.append({
            'Trade_Num': trade_num,
            'Ticker': ticker,
            'Entry_Date': buy_info['entry_date'].strftime('%Y-%m-%d'),
            'NLV_at_Entry': round(buy_info['nlv_at_entry'], 2),
            'Position_Size': round(buy_info['position_size'], 2),
            'Exit_Date': final_date.strftime('%Y-%m-%d'),
            'Gain_Loss_Dollars': round(gain_loss, 2),
            'Gain_Loss_Pct': round(gain_loss_pct, 4),
            'Status': 'Open (Hypothetical Close)'
        })

# Save Trade Log
if trade_log:
    trade_log_df = pd.DataFrame(trade_log)
    
    # FIX: Only divide Gain_Loss_Pct by 100 for Closed trades (not hypothetical)
    trade_log_df.loc[trade_log_df['Status'] == 'Closed', 'Gain_Loss_Pct'] = \
        trade_log_df.loc[trade_log_df['Status'] == 'Closed', 'Gain_Loss_Pct'] / 100
    
    # Save as Excel with formatting
    with pd.ExcelWriter("output/trade_log_summary.xlsx", engine='openpyxl') as writer:
        trade_log_df.to_excel(writer, sheet_name='Trade Log', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Trade Log']
        
        # Format headers
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        
        # Format gain/loss columns with colors
        for row in range(2, len(trade_log) + 2):
            gain_loss_cell = worksheet.cell(row=row, column=7)  # Gain_Loss_Dollars column
            gain_loss_pct_cell = worksheet.cell(row=row, column=8)  # Gain_Loss_Pct column
            status_cell = worksheet.cell(row=row, column=9)  # Status column
            
            if gain_loss_cell.value >= 0:
                gain_loss_cell.font = Font(color="008000")
                gain_loss_pct_cell.font = Font(color="008000")
            else:
                gain_loss_cell.font = Font(color="FF0000")
                gain_loss_pct_cell.font = Font(color="FF0000")
            
            # Highlight hypothetical closes
            if status_cell.value == 'Open (Hypothetical Close)':
                status_cell.font = Font(italic=True, color="0000FF")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"💾 Trade log summary saved to: output/trade_log_summary.xlsx")
    print(f"   Total completed trades: {len([t for t in trade_log if t['Status'] == 'Closed'])}")
    print(f"   Open positions (hypothetically closed): {len([t for t in trade_log if t['Status'] != 'Closed'])}")
    
    # Calculate summary statistics INCLUDING hypothetical closes
    all_gains_losses = [t['Gain_Loss_Dollars'] for t in trade_log]
    total_gains = sum(gl for gl in all_gains_losses if gl > 0)
    total_losses = sum(gl for gl in all_gains_losses if gl < 0)
    winning_trades = len([gl for gl in all_gains_losses if gl > 0])
    losing_trades = len([gl for gl in all_gains_losses if gl < 0])
    
    print(f"\n📈 Trade Statistics (Including Hypothetical Closes):")
    print(f"   Total trades: {len(trade_log)}")
    print(f"   Winning trades: {winning_trades}")
    print(f"   Losing trades: {losing_trades}")
    print(f"   Win rate: {(winning_trades / len(trade_log) * 100):.1f}%")
    print(f"   Total gains: ${total_gains:,.2f}")
    print(f"   Total losses: ${total_losses:,.2f}")
    print(f"   Net P&L (All Trades): ${total_gains + total_losses:,.2f}")
    
    # Show reconciliation
    print(f"\n💰 Portfolio P&L Reconciliation:")
    closed_pnl = sum(t['Gain_Loss_Dollars'] for t in trade_log if t['Status'] == 'Closed')
    hypothetical_pnl = sum(t['Gain_Loss_Dollars'] for t in trade_log if t['Status'] != 'Closed')
    print(f"   Realized P&L (Closed trades): ${closed_pnl:,.2f}")
    print(f"   Unrealized P&L (Hypothetical closes): ${hypothetical_pnl:,.2f}")
    print(f"   Total P&L: ${closed_pnl + hypothetical_pnl:,.2f}")
    print(f"   Portfolio Return: ${final_nlv - INITIAL_CAPITAL:,.2f}")

# Summary statistics
summary = {
    "Initial Capital": INITIAL_CAPITAL,
    "Max Leverage": MAX_LEVERAGE,
    "Position Size %": POSITION_SIZE_PCT * 100,
    "Portfolio Start Date": PORTFOLIO_START_DATE,
    "End Date": END_DATE,
    "Final NLV": round(final_nlv, 2),
    "Total Return $": round(final_nlv - INITIAL_CAPITAL, 2),
    "Total Return %": round(((final_nlv - INITIAL_CAPITAL) / INITIAL_CAPITAL) * 100, 2),
    "Total Trades": len(portfolio.trade_history),
    "Open Positions": len(portfolio.positions),
    "Final Cash": round(portfolio.cash, 2)
}

summary_df = pd.DataFrame([summary])
summary_df.to_excel("output/margin_portfolio_summary.xlsx", index=False)
print(f"💾 Summary saved to: output/margin_portfolio_summary.xlsx")

# Performance metrics by ticker
ticker_performance = {}
for trade in portfolio.trade_history:
    ticker = trade['ticker']
    if ticker not in ticker_performance:
        ticker_performance[ticker] = {'trades': 0, 'pnl': 0}
    ticker_performance[ticker]['trades'] += 1
    if trade['action'] == 'SELL':
        ticker_performance[ticker]['pnl'] += trade.get('pnl', 0)

if ticker_performance:
    perf_df = pd.DataFrame.from_dict(ticker_performance, orient='index')
    perf_df.index.name = 'Ticker'
    perf_df.reset_index(inplace=True)
    perf_df.to_csv("output/ticker_performance.csv", index=False)
    print(f"💾 Ticker performance saved to: output/ticker_performance.csv")

# === Generate Weekly Summary Report ===
print("\n📊 Generating weekly summary report...")

# Function to get week info
def get_week_info(date):
    """Get week number in format YYYY-W##"""
    return f"{date.year}-W{date.isocalendar()[1]:02d}"

# Create fresh tracking variables (ONLY FOR PORTFOLIO PERIOD)
daily_snapshots = []
positions_tracker = {}
cash_tracker = INITIAL_CAPITAL
all_trades = []

# Copy pre-existing positions status
for ticker in pre_existing_positions:
    positions_tracker[ticker] = {'pre_existing': True}

# Track positions day by day (ONLY FROM PORTFOLIO START)
for date in portfolio_dates:
    # Check for sells first
    for ticker in list(positions_tracker.keys()):
        if ticker in ticker_data and date in ticker_data[ticker].index:
            if ticker_data[ticker].loc[date, "signal"] == -1:
                # If it's a pre-existing position, just remove the flag
                if isinstance(positions_tracker.get(ticker), dict) and positions_tracker[ticker].get('pre_existing'):
                    del positions_tracker[ticker]
                    continue
                    
                # Otherwise process the sell
                if ticker in positions_tracker and 'shares' in positions_tracker[ticker]:
                    pos = positions_tracker[ticker]
                    exit_price = ticker_data[ticker].loc[date, "Close"]
                    proceeds = pos['shares'] * exit_price
                    cash_tracker += proceeds
                    
                    # Calculate P&L
                    pnl = proceeds - pos['cost_basis']
                    pnl_pct = (pnl / pos['cost_basis'] * 100) if pos['cost_basis'] > 0 else 0
                    
                    all_trades.append({
                        'date': date,
                        'ticker': ticker,
                        'action': 'SELL',
                        'shares': pos['shares'],
                        'price': exit_price,
                        'pnl': pnl,
                        'pnl_pct': pnl_pct,
                        'entry_date': pos.get('entry_date', date)
                    })
                    
                    del positions_tracker[ticker]
    
    # Check for buys
    buy_candidates = []
    for ticker in ticker_data:
        if ticker not in positions_tracker:
            if date in ticker_data[ticker].index:
                if ticker_data[ticker].loc[date, "signal"] == 1:
                    buy_candidates.append((ticker, ticker_data[ticker].loc[date, "Close"]))
        elif isinstance(positions_tracker.get(ticker), dict) and positions_tracker[ticker].get('pre_existing'):
            continue  # Skip pre-existing positions
    
    if buy_candidates:
        random.shuffle(buy_candidates)
        
        for ticker, price in buy_candidates:
            # Calculate NLV
            nlv = cash_tracker
            for t, p in positions_tracker.items():
                if isinstance(p, dict) and 'shares' in p:
                    if t in ticker_data and date in ticker_data[t].index:
                        nlv += p['shares'] * ticker_data[t].loc[date, "Close"]
            
            # Check buying power
            position_size = nlv * POSITION_SIZE_PCT
            current_positions_value = sum(p.get('cost_basis', 0) for p in positions_tracker.values() 
                                         if isinstance(p, dict) and 'cost_basis' in p)
            max_allowed = nlv * MAX_LEVERAGE
            buying_power = max_allowed - current_positions_value
            
            if buying_power >= position_size:
                shares = int(position_size / price)
                if shares > 0:
                    cost = shares * price
                    cash_tracker -= min(cost, cash_tracker)
                    
                    positions_tracker[ticker] = {
                        'shares': shares,
                        'entry_price': price,
                        'entry_date': date,
                        'cost_basis': cost
                    }
                    
                    all_trades.append({
                        'date': date,
                        'ticker': ticker,
                        'action': 'BUY',
                        'shares': shares,
                        'price': price
                    })
    
    # Create daily snapshot
    position_details = {}
    nlv_end = cash_tracker
    
    for ticker, pos in positions_tracker.items():
        if isinstance(pos, dict) and 'shares' in pos:
            if ticker in ticker_data and date in ticker_data[ticker].index:
                current_price = ticker_data[ticker].loc[date, "Close"]
                current_value = pos['shares'] * current_price
                nlv_end += current_value
                
                pnl = current_value - pos['cost_basis']
                pnl_pct = (pnl / pos['cost_basis'] * 100) if pos['cost_basis'] > 0 else 0
                
                position_details[ticker] = {
                    'shares': pos['shares'],
                    'entry_price': pos['entry_price'],
                    'entry_date': pos['entry_date'],
                    'current_price': current_price,
                    'cost_basis': pos['cost_basis'],
                    'current_value': current_value,
                    'pnl': pnl,
                    'pnl_pct': pnl_pct
                }
    
    total_positions_value = sum(p.get('cost_basis', 0) for p in positions_tracker.values() 
                                if isinstance(p, dict) and 'cost_basis' in p)
    margin_used = max(0, total_positions_value - nlv_end)
    
    daily_snapshots.append({
        'date': date,
        'week': get_week_info(date),
        'nlv': nlv_end,
        'cash': cash_tracker,
        'margin_used': margin_used,
        'num_positions': len(position_details),
        'positions': position_details.copy()
    })

# Group by week and create weekly summaries
weekly_summaries = []
snapshots_df = pd.DataFrame(daily_snapshots)

if not snapshots_df.empty:
    # Group by week and sort chronologically
    week_groups = []
    for week, week_data in snapshots_df.groupby('week'):
        week_data = week_data.sort_values('date')
        first_date = week_data.iloc[0]['date']
        week_groups.append((first_date, week, week_data))
    
    # Sort by actual date
    week_groups.sort(key=lambda x: x[0])
    
    # Process each week
    for first_date, week, week_data in week_groups:
        week_data = week_data.sort_values('date')
        first_day = week_data.iloc[0]
        last_day = week_data.iloc[-1]
        
        # Count trades in this week
        week_start = first_day['date']
        week_end = last_day['date']
        week_trades = [t for t in all_trades 
                      if week_start <= t['date'] <= week_end]
        
        weekly_summary = {
            'Week': week,
            'Start_Date': week_start.strftime('%Y-%m-%d'),
            'End_Date': week_end.strftime('%Y-%m-%d'),
            'Beginning_Balance': round(first_day['nlv'], 2),
            'Ending_Balance': round(last_day['nlv'], 2),
            'Week_Change_Dollars': round(last_day['nlv'] - first_day['nlv'], 2),
            'Week_Change_Pct': round(((last_day['nlv'] - first_day['nlv']) / first_day['nlv']) * 100, 2) if first_day['nlv'] > 0 else 0,
            'Total_Gain_Loss_Dollars': round(last_day['nlv'] - INITIAL_CAPITAL, 2),  # Added this
            'Total_Gain_Loss_Pct': round(((last_day['nlv'] - INITIAL_CAPITAL) / INITIAL_CAPITAL) * 100, 2),  # Added this
            'Cash_Balance': round(last_day['cash'], 2),
            'Margin_Used': round(last_day['margin_used'], 2),
            'Margin_Pct_of_NLV': round((last_day['margin_used'] / last_day['nlv']) * 100, 2) if last_day['nlv'] > 0 else 0,
            'Num_Positions': last_day['num_positions'],
            'Trades_This_Week': len(week_trades),
            'Buys': len([t for t in week_trades if t['action'] == 'BUY']),
            'Sells': len([t for t in week_trades if t['action'] == 'SELL'])
        }
        
        # Add position details
        for i, (ticker, pos) in enumerate(last_day['positions'].items(), 1):
            if i <= 10:
                weekly_summary[f'Pos{i}_Ticker'] = ticker
                weekly_summary[f'Pos{i}_Shares'] = pos['shares']
                weekly_summary[f'Pos{i}_Entry_Date'] = pos.get('entry_date', '')
                weekly_summary[f'Pos{i}_Cost_Basis'] = round(pos['cost_basis'], 2)
                weekly_summary[f'Pos{i}_Current_Value'] = round(pos['current_value'], 2)
                weekly_summary[f'Pos{i}_PnL_Dollars'] = round(pos['pnl'], 2)
                weekly_summary[f'Pos{i}_PnL_Pct'] = round(pos['pnl_pct'], 2)
        
        # Calculate cumulative open P&L
        if last_day['positions']:
            cumulative_open_pnl = sum(pos.get('pnl', 0) for pos in last_day['positions'].values())
        else:
            cumulative_open_pnl = 0
        weekly_summary['Cumulative_Open_PnL'] = round(cumulative_open_pnl, 2)
        
        # Add closed trades
        closed_trades = [t for t in week_trades if t['action'] == 'SELL']
        for i, trade in enumerate(closed_trades[:5], 1):
            weekly_summary[f'Closed{i}_Ticker'] = trade['ticker']
            weekly_summary[f'Closed{i}_Date'] = trade['date'].strftime('%Y-%m-%d')
            weekly_summary[f'Closed{i}_Shares'] = trade['shares']
            weekly_summary[f'Closed{i}_Exit_Price'] = round(trade['price'], 2)
            weekly_summary[f'Closed{i}_PnL'] = round(trade.get('pnl', 0), 2)
            weekly_summary[f'Closed{i}_PnL_Pct'] = round(trade.get('pnl_pct', 0), 2)
        
        weekly_summaries.append(weekly_summary)

# Create Excel with block format
if weekly_summaries:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Portfolio Summary"
    
    # Define styles
    header_font = Font(bold=True, size=12)
    week_header_font = Font(bold=True, size=11, color="FFFFFF")
    week_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    label_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    current_row = 1
    
    # Add report title
    ws.cell(row=current_row, column=1, value="PORTFOLIO WEEKLY SUMMARY REPORT").font = Font(bold=True, size=14)
    current_row += 1
    ws.cell(row=current_row, column=1, value=f"Period: {PORTFOLIO_START_DATE} to {END_DATE}")
    current_row += 1
    ws.cell(row=current_row, column=1, value=f"Initial Capital: ${INITIAL_CAPITAL:,.2f} | Max Leverage: {MAX_LEVERAGE:.1%}")
    current_row += 3
    
    for week_num, week in enumerate(weekly_summaries, 1):
        # Week header
        ws.merge_cells(f'A{current_row}:I{current_row}')
        cell = ws.cell(row=current_row, column=1, 
                      value=f"WEEK {week_num} ({week['Start_Date']} to {week['End_Date']})")
        cell.font = week_header_font
        cell.fill = week_header_fill
        cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Portfolio metrics
        ws.cell(row=current_row, column=1, value="Beginning Balance:").font = label_font
        ws.cell(row=current_row, column=3, value=f"${week['Beginning_Balance']:,.2f}")
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Ending Balance:").font = label_font
        ws.cell(row=current_row, column=3, value=f"${week['Ending_Balance']:,.2f}")
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Week Change:").font = label_font
        change_color = "008000" if week['Week_Change_Dollars'] >= 0 else "FF0000"
        change_cell = ws.cell(row=current_row, column=3, 
                             value=f"${week['Week_Change_Dollars']:,.2f} ({week['Week_Change_Pct']:.2f}%)")
        change_cell.font = Font(color=change_color, bold=True)
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Total Gain/Loss:").font = label_font
        total_color = "008000" if week['Total_Gain_Loss_Dollars'] >= 0 else "FF0000"
        total_cell = ws.cell(row=current_row, column=3, 
                            value=f"${week['Total_Gain_Loss_Dollars']:,.2f} ({week['Total_Gain_Loss_Pct']:.2f}%)")
        total_cell.font = Font(color=total_color, bold=True)
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Cash Balance:").font = label_font
        ws.cell(row=current_row, column=3, value=f"${week['Cash_Balance']:,.2f}")
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Margin Used:").font = label_font
        ws.cell(row=current_row, column=3, value=f"${week['Margin_Used']:,.2f} ({week['Margin_Pct_of_NLV']:.1f}% of NLV)")
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Trades This Week:").font = label_font
        ws.cell(row=current_row, column=3, value=f"{week['Trades_This_Week']} ({week['Buys']} buys, {week['Sells']} sells)")
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Cumulative Open P&L:").font = label_font
        cum_pnl_cell = ws.cell(row=current_row, column=3, value=f"${week.get('Cumulative_Open_PnL', 0):,.2f}")
        cum_pnl_cell.font = Font(color="008000" if week.get('Cumulative_Open_PnL', 0) >= 0 else "FF0000", bold=True)
        current_row += 2
        
        # Active Positions
        if week['Num_Positions'] > 0:
            ws.cell(row=current_row, column=1, value="ACTIVE POSITIONS:").font = label_font
            current_row += 1
            
            # Position headers
            headers = ["Ticker", "Shares", "Entry Date", "Entry Price", "Current Price", "Cost Basis", "Current Value", "P&L $", "P&L %"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = label_font
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.border = border
            current_row += 1
            
            # Position data
            for i in range(1, 11):
                if f'Pos{i}_Ticker' in week and week[f'Pos{i}_Ticker']:
                    ticker = week[f'Pos{i}_Ticker']
                    
                    # Calculate prices
                    entry_price = week[f'Pos{i}_Cost_Basis'] / week[f'Pos{i}_Shares'] if week[f'Pos{i}_Shares'] > 0 else 0
                    current_price = week[f'Pos{i}_Current_Value'] / week[f'Pos{i}_Shares'] if week[f'Pos{i}_Shares'] > 0 else 0
                    
                    # Format entry date
                    entry_date = week.get(f'Pos{i}_Entry_Date', '')
                    if hasattr(entry_date, 'strftime'):
                        entry_date = entry_date.strftime('%Y-%m-%d')
                    elif isinstance(entry_date, str) and len(entry_date) > 10:
                        entry_date = entry_date[:10]
                    
                    ws.cell(row=current_row, column=1, value=ticker).border = border
                    ws.cell(row=current_row, column=2, value=week[f'Pos{i}_Shares']).border = border
                    ws.cell(row=current_row, column=3, value=str(entry_date) if entry_date else '').border = border
                    ws.cell(row=current_row, column=4, value=f"${entry_price:.2f}").border = border
                    ws.cell(row=current_row, column=5, value=f"${current_price:.2f}").border = border
                    ws.cell(row=current_row, column=6, value=f"${week[f'Pos{i}_Cost_Basis']:,.2f}").border = border
                    ws.cell(row=current_row, column=7, value=f"${week[f'Pos{i}_Current_Value']:,.2f}").border = border
                    
                    pnl_cell = ws.cell(row=current_row, column=8, value=f"${week[f'Pos{i}_PnL_Dollars']:,.2f}")
                    pnl_cell.border = border
                    pnl_cell.font = Font(color="008000" if week[f'Pos{i}_PnL_Dollars'] >= 0 else "FF0000")
                    
                    pnl_pct_cell = ws.cell(row=current_row, column=9, value=f"{week[f'Pos{i}_PnL_Pct']:.2f}%")
                    pnl_pct_cell.border = border
                    pnl_pct_cell.font = Font(color="008000" if week[f'Pos{i}_PnL_Pct'] >= 0 else "FF0000")
                    
                    current_row += 1
        else:
            ws.cell(row=current_row, column=1, value="No active positions")
            current_row += 1
        
        current_row += 1
        
        # Closed Trades
        closed_found = False
        for i in range(1, 6):
            if f'Closed{i}_Ticker' in week and week[f'Closed{i}_Ticker']:
                if not closed_found:
                    ws.cell(row=current_row, column=1, value="CLOSED TRADES THIS WEEK:").font = label_font
                    current_row += 1
                    
                    # Closed trade headers
                    closed_headers = ["Ticker", "Exit Date", "Shares", "Exit Price", "P&L $", "P&L %"]
                    for col, header in enumerate(closed_headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=header)
                        cell.font = label_font
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        cell.border = border
                    current_row += 1
                    closed_found = True
                
                # Add closed trade data
                ws.cell(row=current_row, column=1, value=week[f'Closed{i}_Ticker']).border = border
                ws.cell(row=current_row, column=2, value=week[f'Closed{i}_Date']).border = border
                ws.cell(row=current_row, column=3, value=week[f'Closed{i}_Shares']).border = border
                ws.cell(row=current_row, column=4, value=f"${week[f'Closed{i}_Exit_Price']:.2f}").border = border
                
                closed_pnl = week.get(f'Closed{i}_PnL', 0)
                closed_pnl_pct = week.get(f'Closed{i}_PnL_Pct', 0)
                closed_pnl_cell = ws.cell(row=current_row, column=5, value=f"${closed_pnl:,.2f}")
                closed_pnl_cell.border = border
                closed_pnl_cell.font = Font(color="008000" if closed_pnl >= 0 else "FF0000", bold=True)
                
                closed_pnl_pct_cell = ws.cell(row=current_row, column=6, value=f"{closed_pnl_pct:.2f}%")
                closed_pnl_pct_cell.border = border
                closed_pnl_pct_cell.font = Font(color="008000" if closed_pnl_pct >= 0 else "FF0000")
                
                current_row += 1
        
        # Add space between weeks
        current_row += 3
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    try:
        wb.save("output/weekly_portfolio_summary.xlsx")
        print(f"💾 Weekly summary saved to: output/weekly_portfolio_summary.xlsx")
        print(f"   Total weeks analyzed: {len(weekly_summaries)}")
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")
        # Try to save as CSV as backup
        weekly_df = pd.DataFrame(weekly_summaries)
        weekly_df.to_csv("output/weekly_portfolio_summary.csv", index=False)
        print(f"💾 Saved as CSV instead: output/weekly_portfolio_summary.csv")
else:
    print("⚠️ No weekly summaries to save")

print("\n✅ Simulation complete!")